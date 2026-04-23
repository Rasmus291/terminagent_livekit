import os
import re
from typing import Any


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def normalize_phone(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""

    keep_plus = raw.startswith("+")
    digits = re.sub(r"\D", "", raw)
    if raw.startswith("00") and digits.startswith("00"):
        return f"+{digits[2:]}"
    if keep_plus:
        return f"+{digits}"
    return digits


def _pick(row: dict[str, Any], *candidates: str) -> str:
    for candidate in candidates:
        value = row.get(candidate)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def get_contacts_excel_path() -> str:
    return os.getenv("CONTACTS_EXCEL_PATH", "contacts.xlsx")


def load_contacts(excel_path: str | None = None, sheet_name: str | None = None) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Für Excel-Kontaktlisten wird openpyxl benötigt. Bitte installiere das Paket zuerst."
        ) from exc

    workbook_path = excel_path or get_contacts_excel_path()
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(f"Excel-Datei nicht gefunden: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return []

    normalized_headers = [_normalize_header(header) for header in headers]
    contacts: list[dict[str, str]] = []

    for index, values in enumerate(rows, start=2):
        row = {
            normalized_headers[i]: values[i]
            for i in range(min(len(normalized_headers), len(values)))
            if normalized_headers[i]
        }

        contact_id = _pick(row, "id", "contact_id", "kontakt_id") or str(index - 1)
        first_name = _pick(row, "first_name", "vorname")
        last_name = _pick(row, "last_name", "nachname")
        full_name = _pick(row, "name", "full_name", "partner_name", "kontaktname")
        if not full_name:
            full_name = " ".join(part for part in [first_name, last_name] if part).strip()

        phone = _pick(row, "phone", "phone_number", "mobile", "telefon", "telefonnummer", "handy")
        normalized_phone = normalize_phone(phone)
        company = _pick(row, "company", "firma")
        notes = _pick(row, "notes", "notizen", "note")
        salutation = _pick(row, "salutation", "anrede")

        if not full_name and not normalized_phone:
            continue

        contacts.append(
            {
                "contact_id": str(contact_id),
                "name": full_name,
                "first_name": first_name,
                "last_name": last_name,
                "phone": normalized_phone,
                "company": company,
                "notes": notes,
                "salutation": salutation,
            }
        )

    return contacts


def find_contact(contact_id: str | None = None, phone: str | None = None) -> dict[str, str] | None:
    contacts = load_contacts()

    if contact_id:
        contact_id_normalized = str(contact_id).strip()
        for contact in contacts:
            if contact.get("contact_id") == contact_id_normalized:
                return contact

    if phone:
        normalized = normalize_phone(phone)
        for contact in contacts:
            if contact.get("phone") == normalized:
                return contact

    return None